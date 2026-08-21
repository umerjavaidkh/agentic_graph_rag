"""tabular.py — turn CSV / Excel / SQLite into a Neo4j property graph.

Structured data could previously only arrive as a hand-written Cypher script
or a bespoke loader. This reads the shapes people actually have and works out
the graph itself: each table becomes a node label, each row a node, and every
key column that points at another table becomes a relationship.

Inference is deterministic wherever the source can tell us the answer, and
only guesses when it cannot:

* SQLite states its primary and foreign keys, so the relationships are READ,
  not inferred -- no heuristic can beat a declared constraint.
* CSV and Excel carry no constraints, so keys are inferred from column names
  and verified against the data (a claimed key that is not unique is not a
  key, whatever it is called).

An LLM can then improve the NAMES -- "Customer PLACED Order" instead of
"Customer HAS Order" -- but never the structure. Naming is a judgement call
that a model is good at; deciding which 500k rows link to what is a factual
question the data already answers, and a wrong edge there is invisible until
it silently corrupts an answer.

Nothing is written until the caller has seen the plan (see InferredSchema and
the --dry-run path in scripts/load_tabular.py): an inferred relationship that
is wrong looks exactly like a correct one once it is in the graph.
"""
from __future__ import annotations

import csv
import re
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator, Optional

# Rows per write. Large enough that the round trip is amortised, small enough
# that one transaction cannot exhaust the heap on a wide table.
BATCH = 5_000
# Rows read to infer types and uniqueness. The whole file is not needed to
# tell an id column from a description, and reading it twice would double the
# cost of every load.
SAMPLE_ROWS = 2_000

_NON_WORD = re.compile(r"[^0-9a-zA-Z]+")
_ID_SUFFIX = re.compile(r"_?id$", re.I)


# Rows buffered per round trip when streaming a SQL table. Large enough that
# per-batch overhead is negligible, small enough that one batch is not a
# memory event on a wide table.
_SQL_STREAM_BATCH = 1000


def _label_for(table: str) -> str:
    """Table/sheet name -> node label ("olist_order_items" -> "OrderItem")."""
    name = _NON_WORD.sub(" ", table).strip()
    # Common export prefixes carry no meaning in the graph.
    words = [w for w in name.split() if w.lower() not in {"olist", "dataset", "tbl", "table"}]
    words = words or name.split()
    if words and words[-1].lower().endswith("s") and len(words[-1]) > 3:
        words[-1] = words[-1][:-1]  # singular label reads better in Cypher
    return "".join(w[:1].upper() + w[1:] for w in words)


def _rel_for(column: str, target_label: str) -> str:
    """Foreign-key column -> relationship type. Deliberately plain
    (HAS_CUSTOMER); an LLM can propose something better, and structure does
    not depend on the name being good."""
    base = _ID_SUFFIX.sub("", column) or target_label
    return "HAS_" + _NON_WORD.sub("_", base).strip("_").upper()


@dataclass
class Table:
    name: str
    label: str
    columns: list[str]
    primary_key: Optional[str] = None
    # column -> (target table, target column)
    foreign_keys: dict[str, tuple[str, str]] = field(default_factory=dict)
    row_count: Optional[int] = None
    source: str = "csv"

    def describe(self) -> str:
        pk = self.primary_key or "— none found (rows load without identity)"
        fks = ", ".join(f"{c} -> {t}.{k}" for c, (t, k) in self.foreign_keys.items()) or "—"
        rows = f"{self.row_count:,}" if self.row_count is not None else "?"
        return (f"  {self.name} -> (:{self.label})  rows={rows}\n"
                f"      key: {pk}\n      links: {fks}")


@dataclass
class InferredSchema:
    tables: list[Table]
    warnings: list[str] = field(default_factory=list)

    def describe(self) -> str:
        out = [t.describe() for t in self.tables]
        if self.warnings:
            out.append("\n  warnings:")
            out.extend(f"    - {w}" for w in self.warnings)
        return "\n".join(out)


# ── reading ────────────────────────────────────────────────────────────────


def _csv_rows(path: Path) -> Iterator[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        yield from csv.DictReader(fh)


def _sqlite_tables(db: Path) -> list[Table]:
    """Tables, keys and relationships as SQLite itself declares them."""
    con = sqlite3.connect(str(db))
    con.row_factory = sqlite3.Row
    tables: list[Table] = []
    names = [r["name"] for r in con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    )]
    for name in names:
        cols = [r["name"] for r in con.execute(f'PRAGMA table_info("{name}")')]
        pk = next((r["name"] for r in con.execute(f'PRAGMA table_info("{name}")') if r["pk"]), None)
        fks = {
            r["from"]: (r["table"], r["to"] or "rowid")
            for r in con.execute(f'PRAGMA foreign_key_list("{name}")')
        }
        count = con.execute(f'SELECT count(*) AS n FROM "{name}"').fetchone()["n"]
        tables.append(Table(name, _label_for(name), cols, pk, fks, count, source="sqlite"))
    con.close()
    return tables


def _infer_csv_tables(paths: list[Path]) -> InferredSchema:
    """Keys inferred from names, then CHECKED against the data.

    A column called `*_id` is only treated as this table's key if its sampled
    values are actually unique -- `order_id` in an order-items file names the
    order, it does not identify the row, and treating it as identity would
    collapse every line of an order into one node.
    """
    tables: list[Table] = []
    warnings: list[str] = []
    samples: dict[str, list[dict[str, str]]] = {}

    for p in paths:
        rows = []
        for i, row in enumerate(_csv_rows(p)):
            if i >= SAMPLE_ROWS:
                break
            rows.append(row)
        if not rows:
            warnings.append(f"{p.name}: no rows, skipped")
            continue
        samples[p.stem] = rows
        cols = list(rows[0].keys())
        tables.append(Table(p.stem, _label_for(p.stem), cols, source="csv"))

    by_label = {t.label: t for t in tables}
    stem_by_label = {t.label: t.name for t in tables}

    for t in tables:
        rows = samples[t.name]
        for col in t.columns:
            if not _ID_SUFFIX.search(col):
                continue
            values = [r.get(col) for r in rows if r.get(col) not in (None, "")]
            unique = len(set(values)) == len(values) and len(values) == len(rows)
            target = _label_for(_ID_SUFFIX.sub("", col))
            if target in by_label and target != t.label:
                t.foreign_keys[col] = (stem_by_label[target], col)
            elif unique and t.primary_key is None:
                t.primary_key = col
        if t.primary_key is None:
            # Fall back to a key that IS unique even if unnamed as one, so
            # rows still get identity and re-loading is idempotent.
            for col in t.columns:
                vals = [r.get(col) for r in rows]
                if len(set(vals)) == len(vals) and all(v not in (None, "") for v in vals):
                    t.primary_key = col
                    break
        if t.primary_key is None:
            warnings.append(
                f"{t.name}: no unique column found; rows will be created without identity "
                "and a second load would duplicate them"
            )

    # A second pass, once every table's key is known: a column whose NAME
    # matches another table's primary key is a foreign key even when it
    # breaks the "_id" convention. Verified on a real export --
    # products.product_category_name points at the category table's key and
    # the suffix rule missed it entirely, which would have left every product
    # with no category and made "revenue by category" unanswerable.
    keys = {t.primary_key: t for t in tables if t.primary_key}
    for t in tables:
        for col in t.columns:
            owner = keys.get(col)
            if owner is not None and owner.name != t.name and col not in t.foreign_keys:
                t.foreign_keys[col] = (owner.name, col)

    return InferredSchema(tables, warnings)


# ── SQL databases ──────────────────────────────────────────────────────────
#
# A live database is the one source that already KNOWS its schema. CSVs and
# spreadsheets force us to guess relationships from column names; a database
# declares them, so reflection gives real foreign keys instead of inference.
#
# SQLAlchemy is imported lazily throughout. It is an optional dependency -- a
# deployment that only ingests files should not have to install a database
# toolkit, and this module must stay importable without it.

# Anything that is not an existing path is treated as a connection URL.
# Checked in that order deliberately: a relative path like "data/shop.db"
# contains no "://" but is still a file, and probing the filesystem first
# means a typo'd URL fails with a connection error rather than silently
# being read as a missing file.
def _as_sql_url(source: "str | Path") -> Optional[str]:
    """The connection URL, if `source` names a database rather than a file."""
    if isinstance(source, Path):
        return None
    text = str(source)
    return text if "://" in text and not Path(text).exists() else None


def _safe_url(url: str) -> str:
    """The URL with its password removed, for logs and provenance stamps.

    A connection URL routinely carries credentials, and `_source` is written
    onto every node this loader creates. Stamping the raw URL would put the
    database password in the graph, where it is readable by anyone who can
    query it and survives in backups.
    """
    from sqlalchemy.engine import make_url

    try:
        return make_url(url).render_as_string(hide_password=True)
    except Exception:
        # Never let a malformed URL leak by falling back to the raw string.
        return url.split("://", 1)[0] + "://<unparsed>"


def _sql_engine(url: str):
    from sqlalchemy import create_engine

    # pool_pre_ping costs one round trip per checkout and buys immunity to
    # connections dropped by an idle timeout, which is the normal failure
    # mode for a long ingest against a managed database.
    return create_engine(url, pool_pre_ping=True)


def _sqlalchemy_tables(url: str) -> InferredSchema:
    """Reflect a database into the same Table shape the file loaders produce.

    Uses DECLARED foreign keys. That is the whole reason to prefer a database
    source: `_infer_csv_tables` has to guess a relationship from a column
    called `customer_id`, and guesses wrongly on any column that merely looks
    like a key. Reflection reports what the schema actually says.
    """
    from sqlalchemy import MetaData, inspect

    engine = _sql_engine(url)
    meta = MetaData()
    try:
        meta.reflect(bind=engine)
        inspector = inspect(engine)
        tables: list[Table] = []
        warnings: list[str] = []
        for name, tbl in sorted(meta.tables.items()):
            pk_cols = [c.name for c in tbl.primary_key.columns]
            if len(pk_cols) > 1:
                # A composite key has no single property to MERGE on, so the
                # rows would load without identity and duplicate on re-run.
                warnings.append(
                    f"{name}: composite primary key ({', '.join(pk_cols)}) is not "
                    "supported -- rows will load without identity"
                )
            fks: dict[str, tuple[str, str]] = {}
            for fk in tbl.foreign_keys:
                fks[fk.parent.name] = (fk.column.table.name, fk.column.name)
            tables.append(Table(
                name=name,
                label=_label_for(name),
                columns=[c.name for c in tbl.columns],
                primary_key=pk_cols[0] if len(pk_cols) == 1 else None,
                foreign_keys=fks,
                row_count=None,      # counting every table up front costs a
                                     # full scan each; the loader reports real
                                     # counts as it writes.
                source="sql",
            ))
        if tables and not any(t.foreign_keys for t in tables):
            warnings.append(
                "no foreign keys declared in this database, so no relationships were "
                "found -- the graph will be disconnected islands unless keys are added"
            )
        if not tables:
            warnings.append(f"{_safe_url(url)}: no tables found")
        _ = inspector  # reflection already used it; kept for clarity of intent
        return InferredSchema(tables, warnings)
    finally:
        engine.dispose()


def _sql_rows(url: str, table: Table) -> Iterator[dict[str, Any]]:
    """Stream one table's rows without materialising it.

    `stream_results` asks the driver for a server-side cursor and `yield_per`
    bounds what is buffered, so a table of any size costs the same memory
    here. The connection stays open for the length of the walk -- that is the
    trade being made, and it is the right one: the alternative is holding
    millions of rows in the process instead.
    """
    from sqlalchemy import MetaData, Table as SATable, select

    engine = _sql_engine(url)
    try:
        meta = MetaData()
        sa_table = SATable(table.name, meta, autoload_with=engine)
        with engine.connect().execution_options(
            stream_results=True, yield_per=_SQL_STREAM_BATCH
        ) as conn:
            for row in conn.execute(select(sa_table)):
                yield dict(row._mapping)
    finally:
        engine.dispose()


def infer_schema(source: "str | Path") -> InferredSchema:
    """Read a directory of CSVs, an .xlsx workbook, a SQLite file, or a database URL."""
    url = _as_sql_url(source)
    if url:
        return _sqlalchemy_tables(url)
    source = Path(source)
    if source.is_dir():
        csvs = sorted(source.glob("*.csv"))
        if not csvs:
            raise ValueError(f"no .csv files in {source}")
        return _infer_csv_tables(csvs)
    if source.suffix.lower() in {".db", ".sqlite", ".sqlite3"}:
        tables = _sqlite_tables(source)
        warn = [] if tables else [f"{source.name}: no tables found"]
        # A SQLite file with no declared foreign keys tells us nothing more
        # than a pile of CSVs would, and the caller should know that before
        # trusting the plan.
        if tables and not any(t.foreign_keys for t in tables):
            warn.append(
                "no foreign keys declared in this database, so no relationships were "
                "found -- the graph will be disconnected islands unless keys are added"
            )
        return InferredSchema(tables, warn)
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        return _infer_excel(source)
    raise ValueError(f"unsupported source: {source}")


def _infer_excel(path: Path) -> InferredSchema:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise ValueError("reading .xlsx needs openpyxl (pip install openpyxl)") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    tmp_tables: list[Table] = []
    samples: dict[str, list[dict[str, str]]] = {}
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            continue
        cols = [str(h) for h in header if h is not None]
        rows = []
        for i, values in enumerate(it):
            if i >= SAMPLE_ROWS:
                break
            rows.append({c: ("" if v is None else str(v)) for c, v in zip(cols, values)})
        if not rows:
            continue
        samples[ws.title] = rows
        tmp_tables.append(Table(ws.title, _label_for(ws.title), cols, source="excel"))
    wb.close()

    # Same key/relationship inference as CSV -- a worksheet and a CSV file are
    # the same thing once read.
    schema = InferredSchema(tmp_tables)
    by_label = {t.label: t.name for t in tmp_tables}
    for t in tmp_tables:
        rows = samples[t.name]
        for col in t.columns:
            if not _ID_SUFFIX.search(col):
                continue
            target = _label_for(_ID_SUFFIX.sub("", col))
            values = [r.get(col) for r in rows if r.get(col) not in (None, "")]
            if target in by_label and by_label[target] != t.name:
                t.foreign_keys[col] = (by_label[target], col)
            elif t.primary_key is None and len(set(values)) == len(values) == len(rows):
                t.primary_key = col
    return schema


# ── loading ────────────────────────────────────────────────────────────────


def _rows_for(source: "str | Path", table: Table) -> Iterator[dict[str, Any]]:
    if table.source == "sql":
        # `yield from`, not `return`. This function contains yields elsewhere,
        # which makes it a generator function -- so `return <iterator>` sets a
        # StopIteration value instead of handing back the rows, and every SQL
        # table would load as zero rows with no error anywhere.
        yield from _sql_rows(str(source), table)
        return
    source = Path(source)
    if table.source == "sqlite":
        con = sqlite3.connect(str(source))
        con.row_factory = sqlite3.Row
        try:
            for r in con.execute(f'SELECT * FROM "{table.name}"'):
                yield {k: r[k] for k in r.keys()}
        finally:
            con.close()
    elif table.source == "excel":
        from openpyxl import load_workbook

        wb = load_workbook(source, read_only=True, data_only=True)
        ws = wb[table.name]
        it = ws.iter_rows(values_only=True)
        header = [str(h) for h in (next(it, None) or []) if h is not None]
        for values in it:
            yield {c: ("" if v is None else v) for c, v in zip(header, values)}
        wb.close()
    else:
        yield from _csv_rows(source / f"{table.name}.csv")


def source_tag(source: "str | Path") -> str:
    """Provenance stamp written on every node this loader creates.

    Labels are inferred from table names, so two unrelated sources collide
    constantly -- a two-row test fixture with a `products` table infers the
    same :Product label as a 33k-row production dataset. Without a stamp,
    clearing the fixture deletes the real data and the loss is silent.
    Verified the hard way: exactly that wiped 36,046 rows here.
    """
    url = _as_sql_url(source)
    if url:
        # Credentials stripped: this string is written onto every node.
        return f"tabular:{_safe_url(url)}"
    return f"tabular:{Path(source).name}"


def load_schema(session, source: "str | Path", schema: InferredSchema) -> dict[str, int]:
    """Create nodes then relationships, returning per-table row counts.

    Nodes for every table are written BEFORE any relationship, because a
    relationship whose other end has not been created yet silently matches
    nothing -- the load would appear to succeed and produce an unlinked graph.
    """
    counts: dict[str, int] = {}

    for t in schema.tables:
        if t.primary_key:
            session.run(
                f"CREATE CONSTRAINT tab_{t.label.lower()} IF NOT EXISTS "
                f"FOR (n:{t.label}) REQUIRE n.{t.primary_key} IS UNIQUE"
            )

    tag = source_tag(source)
    for t in schema.tables:
        merge = (f"MERGE (n:{t.label} {{{t.primary_key}: row.{t.primary_key}}}) SET n += row"
                 if t.primary_key else f"CREATE (n:{t.label}) SET n += row")
        cypher = f"UNWIND $rows AS row {merge} SET n._source = $tag"
        counts[t.name] = _write(session, cypher, _rows_for(source, t), tag=tag)

    for t in schema.tables:
        for col, (target_table, target_col) in t.foreign_keys.items():
            target = next((x for x in schema.tables if x.name == target_table), None)
            if target is None or not target.primary_key or not t.primary_key:
                continue
            cypher = (
                f"UNWIND $rows AS row "
                f"MATCH (a:{t.label} {{{t.primary_key}: row.src}}) "
                f"MATCH (b:{target.label} {{{target.primary_key}: row.dst}}) "
                f"MERGE (a)-[:{_rel_for(col, target.label)}]->(b)"
            )
            pairs = (
                {"src": r.get(t.primary_key), "dst": r.get(col)}
                for r in _rows_for(source, t)
                if r.get(col) not in (None, "")
            )
            _write(session, cypher, pairs)
    return counts


def _write(session, cypher: str, rows: Iterator[dict[str, Any]], tag: Optional[str] = None) -> int:
    batch: list[dict[str, Any]] = []
    total = 0
    params: dict[str, Any] = {"tag": tag} if tag else {}
    for row in rows:
        batch.append(row)
        if len(batch) >= BATCH:
            session.run(cypher, rows=batch, **params)
            total += len(batch)
            batch = []
    if batch:
        session.run(cypher, rows=batch, **params)
        total += len(batch)
    return total
